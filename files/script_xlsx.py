from openpyxl import load_workbook


workbook = load_workbook('files_to_zip/1518_SEPTEMBER.xlsx')
sheet = workbook.active
print(sheet.cell(row=5, column=5).value)

for x in sheet.columns:
    print(x)
