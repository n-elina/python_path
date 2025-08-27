import os
from zipfile import ZipFile, ZIP_DEFLATED
import pytest
from pypdf import PdfReader
from openpyxl import load_workbook
from files.script_os import FILES_DIR


ZIP_PATH = os.path.join(FILES_DIR, 'archive.zip')

# @pytest.fixture
def test_create_archive():
    files_to_zip = [
        os.path.join(FILES_DIR, '40348747.pdf'),
        os.path.join(FILES_DIR, '1518_SEPTEMBER.xlsx'),
        os.path.join(FILES_DIR, 'sample.csv')
    ]

    # zip_path = os.path.join(FILES_DIR, 'archive.zip')

    with ZipFile(ZIP_PATH, mode='w', compression=ZIP_DEFLATED) as zip_create:
        for file in files_to_zip:
            zip_create.write(file, os.path.basename(file))
    assert os.path.exists(ZIP_PATH)
    # return zip_path


def test_pdf_file_in_archive():
    with ZipFile(ZIP_PATH) as zip_archive:
        pdf = zip_archive.open('40348747.pdf')
        reader = PdfReader(pdf)
        text = reader.pages[0].extract_text()
    assert 'Н е л ю б о в  М и к о л а' in text


def test_xlsx_file_in_archive():
    with ZipFile(ZIP_PATH) as zip_archive:
        xlsx = zip_archive.open('1518_SEPTEMBER.xlsx')
        workbook = load_workbook(xlsx).active
        project_name = workbook.cell(row=5, column=3).value
    assert 'WordScape' in project_name


def test_csv_file_in_archive():
    with ZipFile(ZIP_PATH) as zip_archive:
        csv = zip_archive.open('sample.csv')
        reader = csv.read().decode('utf-8')
    assert 'Alice,30,New York' in reader
